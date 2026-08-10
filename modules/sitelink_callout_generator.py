# PowerShell 运行（请先进入项目根目录）：
# & ".\.venv\Scripts\python.exe" -m modules.sitelink_callout_generator

from __future__ import annotations

import argparse
import json
import re

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from modules.llm_client import call_llm
from modules.prompt_loader import load_prompt
from modules.website_url_extractor import resolve_website_urls_path


SITELINK_TITLE_LIMIT = 30
SITELINK_DESCRIPTION_LIMIT = 60
CALLOUT_LIMIT = 25
CALLOUT_DESKTOP_TOTAL_LIMIT = 132
CALLOUT_MOBILE_TOTAL_LIMIT = 66

FORBIDDEN_CHARACTERS = {"!", "?", "[", "]"}


def get_project_root() -> Path:
    return Path(__file__).resolve().parent.parent


def get_output_dir(project_name: str) -> Path:
    return get_project_root() / "outputs" / project_name


def read_required_text(path: Path, label: str) -> str:
    if not path.exists():
        raise FileNotFoundError(
            f"找不到{label}：{path}"
        )

    text = path.read_text(encoding="utf-8")

    if not text.strip():
        raise ValueError(
            f"{label}为空：{path}"
        )

    return text


def load_website_json(project_name: str) -> dict:
    path = get_output_dir(project_name) / "website_pages.json"

    if not path.exists():
        raise FileNotFoundError(
            f"找不到网页读取结果：{path}\n"
            f"请先运行 modules.web_reader。"
        )

    try:
        data = json.loads(
            path.read_text(encoding="utf-8")
        )
    except json.JSONDecodeError as error:
        raise ValueError(
            f"website_pages.json 不是合法JSON：{error}"
        ) from error

    pages = data.get("pages")

    if not isinstance(pages, list) or not pages:
        raise ValueError(
            "website_pages.json 中没有成功读取的网页。"
        )

    return data


def clean_one_line(text: str | None) -> str:
    return re.sub(
        r"\s+",
        " ",
        str(text or ""),
    ).strip()


def build_page_inventory(
    website_data: dict,
    max_pages: int = 25,
    excerpt_chars: int = 1200,
) -> tuple[str, dict]:
    """
    把大型 website_pages.json 压缩成适合发送给AI的页面清单。

    AI只看到页面编号，不直接决定URL。
    """
    pages = website_data.get("pages", [])

    # 优先使用抓取深度较浅的页面
    pages = sorted(
        pages,
        key=lambda page: (
            page.get("depth", 999),
            page.get("url", ""),
        ),
    )

    inventory_sections = []
    page_map = {}

    for index, page in enumerate(
        pages[:max_pages],
        start=1,
    ):
        page_id = f"P{index:03d}"

        url = clean_one_line(page.get("url"))
        title = clean_one_line(page.get("title"))
        meta_description = clean_one_line(
            page.get("meta_description")
        )
        page_text = clean_one_line(page.get("text"))

        headings = page.get("headings", [])
        heading_texts = []

        if isinstance(headings, list):
            for heading in headings[:12]:
                if not isinstance(heading, dict):
                    continue

                level = clean_one_line(
                    heading.get("level")
                ).upper()
                text = clean_one_line(
                    heading.get("text")
                )

                if text:
                    heading_texts.append(
                        f"{level}: {text}"
                    )

        page_map[page_id] = {
            "page_id": page_id,
            "url": url,
            "title": title,
            "meta_description": meta_description,
            "headings": heading_texts,
            "text_excerpt": page_text[:excerpt_chars],
        }

        inventory_sections.extend([
            f"## {page_id}",
            f"URL: {url}",
            f"Title: {title or '无'}",
            (
                "Meta Description: "
                f"{meta_description or '无'}"
            ),
            (
                "Headings: "
                + (
                    " | ".join(heading_texts)
                    if heading_texts
                    else "无"
                )
            ),
            (
                "Content Excerpt: "
                f"{page_text[:excerpt_chars] or '无'}"
            ),
            "",
        ])

    if not page_map:
        raise ValueError(
            "没有可提供给AI的网站页面。"
        )

    return "\n".join(inventory_sections), page_map


def build_prompt(
    project_brief: str,
    website_inventory: str,
) -> str:
    prompt = load_prompt("sitelink_callout")

    replacements = {
        "project_brief": project_brief,
        "website_pages": website_inventory,
    }

    for key, value in replacements.items():
        prompt = prompt.replace(
            "{{" + key + "}}",
            value,
        )

    return prompt


def build_callout_only_prompt(project_brief: str) -> str:
    prompt = load_prompt("callout_only")
    return prompt.replace("{{project_brief}}", project_brief)


def extract_json_from_response(text: str) -> dict:
    """
    兼容模型偶尔返回 ```json 代码块的情况。
    """
    cleaned = text.strip()

    cleaned = re.sub(
        r"^```(?:json)?\s*",
        "",
        cleaned,
        flags=re.IGNORECASE,
    )
    cleaned = re.sub(
        r"\s*```$",
        "",
        cleaned,
    )

    first_brace = cleaned.find("{")
    last_brace = cleaned.rfind("}")

    if first_brace == -1 or last_brace == -1:
        raise ValueError(
            "AI没有返回JSON对象。"
        )

    json_text = cleaned[
        first_brace:last_brace + 1
    ]

    try:
        return json.loads(json_text)
    except json.JSONDecodeError as error:
        raise ValueError(
            f"AI返回的JSON无法解析：{error}"
        ) from error


def contains_forbidden_characters(text: str) -> bool:
    return any(
        character in text
        for character in FORBIDDEN_CHARACTERS
    )


def normalize_for_duplicate_check(text: str) -> str:
    return clean_one_line(text).lower()


def validate_payload(
    payload: dict,
    page_map: dict,
    sitelinks_required: bool = True,
) -> list[str]:
    """
    返回所有校验错误。

    没有错误时返回空列表。
    """
    errors = []

    sitelinks = payload.get("sitelinks")
    callouts = payload.get("callouts")

    if sitelinks is None and not sitelinks_required:
        sitelinks = []

    if not isinstance(sitelinks, list):
        return ["sitelinks 必须是数组。"]

    if not isinstance(callouts, list):
        return ["callouts 必须是数组。"]

    if sitelinks_required and len(sitelinks) != 8:
        errors.append(
            f"Sitelink必须正好8条，实际为{len(sitelinks)}条。"
        )
    elif not sitelinks_required and sitelinks:
        errors.append("无网站模式不得生成Sitelink。")

    sitelink_titles = set()
    sitelink_descriptions = set()

    for index, item in enumerate(
        sitelinks,
        start=1,
    ):
        if not isinstance(item, dict):
            errors.append(
                f"Sitelink {index} 不是JSON对象。"
            )
            continue

        page_id = clean_one_line(
            item.get("page_id")
        )
        title_ru = clean_one_line(
            item.get("title_ru")
        )
        title_zh = clean_one_line(
            item.get("title_zh")
        )
        description_ru = clean_one_line(
            item.get("description_ru")
        )
        description_zh = clean_one_line(
            item.get("description_zh")
        )

        if page_id not in page_map:
            errors.append(
                f"Sitelink {index} 使用了不存在的page_id："
                f"{page_id}"
            )

        if not title_ru:
            errors.append(
                f"Sitelink {index} 缺少俄语标题。"
            )
        elif len(title_ru) > SITELINK_TITLE_LIMIT:
            errors.append(
                f"Sitelink {index} 标题超过"
                f"{SITELINK_TITLE_LIMIT}字符："
                f"{len(title_ru)}字符。"
            )

        if not title_zh:
            errors.append(
                f"Sitelink {index} 缺少标题中文翻译。"
            )

        if not description_ru:
            errors.append(
                f"Sitelink {index} 缺少俄语描述。"
            )
        elif (
            len(description_ru)
            > SITELINK_DESCRIPTION_LIMIT
        ):
            errors.append(
                f"Sitelink {index} 描述超过"
                f"{SITELINK_DESCRIPTION_LIMIT}字符："
                f"{len(description_ru)}字符。"
            )

        if not description_zh:
            errors.append(
                f"Sitelink {index} 缺少描述中文翻译。"
            )

        if contains_forbidden_characters(title_ru):
            errors.append(
                f"Sitelink {index} 标题包含禁用字符。"
            )

        if contains_forbidden_characters(
            description_ru
        ):
            errors.append(
                f"Sitelink {index} 描述包含禁用字符。"
            )

        normalized_title = (
            normalize_for_duplicate_check(title_ru)
        )
        normalized_description = (
            normalize_for_duplicate_check(
                description_ru
            )
        )

        if normalized_title:
            if normalized_title in sitelink_titles:
                errors.append(
                    f"Sitelink {index} 标题重复："
                    f"{title_ru}"
                )
            sitelink_titles.add(normalized_title)

        if normalized_description:
            if (
                normalized_description
                in sitelink_descriptions
            ):
                errors.append(
                    f"Sitelink {index} 描述重复："
                    f"{description_ru}"
                )
            sitelink_descriptions.add(
                normalized_description
            )

    if len(callouts) < 8:
        errors.append(
            f"Callout至少需要8条，实际为{len(callouts)}条。"
        )

    callout_texts = set()

    for index, item in enumerate(
        callouts,
        start=1,
    ):
        if not isinstance(item, dict):
            errors.append(
                f"Callout {index} 不是JSON对象。"
            )
            continue

        text_ru = clean_one_line(
            item.get("text_ru")
        )
        text_zh = clean_one_line(
            item.get("text_zh")
        )

        if not text_ru:
            errors.append(
                f"Callout {index} 缺少俄语内容。"
            )
        elif len(text_ru) > CALLOUT_LIMIT:
            errors.append(
                f"Callout {index} 超过"
                f"{CALLOUT_LIMIT}字符："
                f"{len(text_ru)}字符。"
            )

        if not text_zh:
            errors.append(
                f"Callout {index} 缺少中文翻译。"
            )

        if contains_forbidden_characters(text_ru):
            errors.append(
                f"Callout {index} 包含禁用字符。"
            )

        normalized_text = (
            normalize_for_duplicate_check(text_ru)
        )

        if normalized_text:
            if normalized_text in callout_texts:
                errors.append(
                    f"Callout {index} 内容重复："
                    f"{text_ru}"
                )
            callout_texts.add(normalized_text)

    if len(callouts) >= 8:
        desktop_total = sum(
            len(
                clean_one_line(
                    item.get("text_ru")
                )
            )
            for item in callouts[:8]
            if isinstance(item, dict)
        )

        if (
            desktop_total
            > CALLOUT_DESKTOP_TOTAL_LIMIT
        ):
            errors.append(
                f"前8条Callout合计{desktop_total}字符，"
                f"超过桌面端"
                f"{CALLOUT_DESKTOP_TOTAL_LIMIT}字符限制。"
            )

    if len(callouts) >= 4:
        mobile_total = sum(
            len(
                clean_one_line(
                    item.get("text_ru")
                )
            )
            for item in callouts[:4]
            if isinstance(item, dict)
        )

        if (
            mobile_total
            > CALLOUT_MOBILE_TOTAL_LIMIT
        ):
            errors.append(
                f"前4条Callout合计{mobile_total}字符，"
                f"超过移动端"
                f"{CALLOUT_MOBILE_TOTAL_LIMIT}字符限制。"
            )

    return errors


def enrich_payload(
    payload: dict,
    page_map: dict,
    result_meta: dict,
    sitelinks_status: str = "generated",
) -> dict:
    """
    添加真实URL、字符数和模型信息。
    """
    enriched_sitelinks = []

    for index, item in enumerate(
        payload["sitelinks"],
        start=1,
    ):
        page_id = clean_one_line(
            item.get("page_id")
        )
        page = page_map[page_id]

        title_ru = clean_one_line(
            item.get("title_ru")
        )
        description_ru = clean_one_line(
            item.get("description_ru")
        )

        enriched_sitelinks.append({
            "index": index,
            "page_id": page_id,
            "url": page["url"],
            "source_page_title": page["title"],
            "title_ru": title_ru,
            "title_zh": clean_one_line(
                item.get("title_zh")
            ),
            "title_char_count": len(title_ru),
            "title_char_limit": SITELINK_TITLE_LIMIT,
            "description_ru": description_ru,
            "description_zh": clean_one_line(
                item.get("description_zh")
            ),
            "description_char_count": len(
                description_ru
            ),
            "description_char_limit": (
                SITELINK_DESCRIPTION_LIMIT
            ),
            "status": "OK",
        })

    enriched_callouts = []

    for index, item in enumerate(
        payload["callouts"],
        start=1,
    ):
        text_ru = clean_one_line(
            item.get("text_ru")
        )

        if index <= 4:
            recommendation = "移动端优先 + 桌面端主选"
        elif index <= 8:
            recommendation = "桌面端主选"
        else:
            recommendation = "备用候选"

        enriched_callouts.append({
            "index": index,
            "text_ru": text_ru,
            "text_zh": clean_one_line(
                item.get("text_zh")
            ),
            "char_count": len(text_ru),
            "char_limit": CALLOUT_LIMIT,
            "recommendation": recommendation,
            "status": "OK",
        })

    return {
        "sitelinks": enriched_sitelinks,
        "sitelinks_status": sitelinks_status,
        "callouts": enriched_callouts,
        "summary": {
            "sitelink_count": len(
                enriched_sitelinks
            ),
            "callout_count": len(
                enriched_callouts
            ),
            "desktop_primary_total_chars": sum(
                item["char_count"]
                for item in enriched_callouts[:8]
            ),
            "mobile_primary_total_chars": sum(
                item["char_count"]
                for item in enriched_callouts[:4]
            ),
        },
        "model": {
            "actual_model": result_meta.get(
                "actual_model"
            ),
            "provider": result_meta.get(
                "provider"
            ),
            "finish_reason": result_meta.get(
                "finish_reason"
            ),
            "usage": result_meta.get(
                "usage",
                {},
            ),
        },
    }


def write_json(path: Path, data: dict) -> None:
    path.write_text(
        json.dumps(
            data,
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def set_cell_style(
    cell,
    fill=None,
    bold=False,
    font_color="000000",
    horizontal="center",
):
    thin = Side(
        style="thin",
        color="808080",
    )

    cell.border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin,
    )
    cell.alignment = Alignment(
        horizontal=horizontal,
        vertical="center",
        wrap_text=True,
    )
    cell.font = Font(
        name="Microsoft YaHei",
        size=11,
        bold=bold,
        color=font_color,
    )

    if fill:
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=fill,
        )


def write_xlsx(path: Path, data: dict) -> None:
    workbook = Workbook()
    # Workbook() 会创建一个默认工作表。直接从 worksheets 读取后，
    # Pylance 能确定这里一定是 Worksheet，而不是可能为 None。
    sheet = workbook.worksheets[0]
    sheet.title = "sitelink+callouts"
    sheet.sheet_view.showGridLines = False

    header_fill = "95B3D7"
    section_fill = "DCE6F1"
    warning_fill = "FFF2CC"

    row = 1

    if data.get("sitelinks_status") == "skipped_no_website":
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        notice_cell = sheet.cell(
            row,
            1,
            "客户资料未提供网站，Sitelink未生成；以下仅输出Callout候选。",
        )
        set_cell_style(
            notice_cell,
            fill=warning_fill,
            bold=True,
        )
        row += 2

    sheet.cell(row, 1, "项目")
    sheet.cell(row, 2, "内容-俄语")
    sheet.cell(row, 3, "字符数")
    sheet.cell(row, 4, "内容-中文")
    sheet.cell(row, 5, "状态/来源")

    for column in range(1, 6):
        set_cell_style(
            sheet.cell(row, column),
            fill=header_fill,
            bold=True,
        )

    row += 1

    for sitelink in data["sitelinks"]:
        index = sitelink["index"]

        sheet.cell(row, 1, f"Sitelink {index}")
        sheet.cell(row, 2, f"Sitelink {index}")
        sheet.cell(row, 4, f"Sitelink {index}")
        sheet.cell(
            row,
            5,
            (
                f"{sitelink['page_id']} | "
                f"{sitelink['source_page_title']}"
            ),
        )

        for column in range(1, 6):
            set_cell_style(
                sheet.cell(row, column),
                fill=section_fill,
            )

        row += 1

        sheet.cell(
            row,
            1,
            f"Title（{SITELINK_TITLE_LIMIT}字符）",
        )
        sheet.cell(row, 2, sitelink["title_ru"])
        sheet.cell(
            row,
            3,
            sitelink["title_char_count"],
        )
        sheet.cell(row, 4, sitelink["title_zh"])
        sheet.cell(row, 5, sitelink["status"])

        row += 1

        sheet.cell(
            row,
            1,
            (
                "Description"
                f"（{SITELINK_DESCRIPTION_LIMIT}字符）"
            ),
        )
        sheet.cell(
            row,
            2,
            sitelink["description_ru"],
        )
        sheet.cell(
            row,
            3,
            sitelink["description_char_count"],
        )
        sheet.cell(
            row,
            4,
            sitelink["description_zh"],
        )
        sheet.cell(row, 5, sitelink["status"])

        row += 1

        sheet.cell(row, 1, "URL")
        sheet.cell(row, 2, sitelink["url"])
        sheet.cell(row, 4, sitelink["url"])
        sheet.cell(
            row,
            5,
            "已从website_pages.json映射",
        )

        sheet.cell(row, 2).hyperlink = sitelink["url"]
        sheet.cell(row, 4).hyperlink = sitelink["url"]
        sheet.cell(row, 2).style = "Hyperlink"
        sheet.cell(row, 4).style = "Hyperlink"

        for current_row in range(row - 2, row + 1):
            for column in range(1, 6):
                set_cell_style(
                    sheet.cell(current_row, column),
                )

        row += 1

    row += 2

    sheet.cell(row, 1, "Callout")
    sheet.cell(row, 2, "内容-俄语")
    sheet.cell(row, 3, "字符数")
    sheet.cell(row, 4, "内容-中文")
    sheet.cell(row, 5, "投放建议")

    for column in range(1, 6):
        set_cell_style(
            sheet.cell(row, column),
            fill=header_fill,
            bold=True,
        )

    row += 1

    for callout in data["callouts"]:
        sheet.cell(
            row,
            1,
            (
                f"标注{callout['index']}"
                f"（{CALLOUT_LIMIT}字符）"
            ),
        )
        sheet.cell(row, 2, callout["text_ru"])
        sheet.cell(row, 3, callout["char_count"])
        sheet.cell(row, 4, callout["text_zh"])
        sheet.cell(
            row,
            5,
            callout["recommendation"],
        )

        fill = (
            warning_fill
            if callout["index"] > 8
            else None
        )

        for column in range(1, 6):
            set_cell_style(
                sheet.cell(row, column),
                fill=fill,
            )

        row += 1

    row += 1

    summary = data["summary"]

    sheet.cell(
        row,
        1,
        "桌面端前8条合计",
    )
    sheet.cell(
        row,
        3,
        summary["desktop_primary_total_chars"],
    )
    sheet.cell(
        row,
        5,
        f"限制≤{CALLOUT_DESKTOP_TOTAL_LIMIT}",
    )

    row += 1

    sheet.cell(
        row,
        1,
        "移动端前4条合计",
    )
    sheet.cell(
        row,
        3,
        summary["mobile_primary_total_chars"],
    )
    sheet.cell(
        row,
        5,
        f"限制≤{CALLOUT_MOBILE_TOTAL_LIMIT}",
    )

    for current_row in range(row - 1, row + 1):
        for column in range(1, 6):
            set_cell_style(
                sheet.cell(current_row, column),
                fill=section_fill,
                bold=True,
            )

    column_widths = {
        1: 25,
        2: 70,
        3: 12,
        4: 45,
        5: 42,
    }

    for column, width in column_widths.items():
        sheet.column_dimensions[
            get_column_letter(column)
        ].width = width

    for current_row in range(
        1,
        sheet.max_row + 1,
    ):
        sheet.row_dimensions[current_row].height = 24

    sheet.freeze_panes = "A2"

    workbook.save(path)


def generate_sitelink_callouts(
    project_name: str,
    provider=None,
    model=None,
    temperature: float = 0.5,
    max_tokens: int = 5000,
    use_premium=None,
    max_page_candidates: int = 25,
    excerpt_chars: int = 1200,
) -> dict:
    output_dir = get_output_dir(project_name)

    project_brief_path = (
        output_dir / "project_brief.md"
    )

    project_brief = read_required_text(
        project_brief_path,
        "项目资料分析 project_brief.md",
    )

    website_url_path, website_url_source = resolve_website_urls_path(
        project_root=get_project_root(),
        project_name=project_name,
    )

    has_website = website_url_path is not None

    if has_website:
        if website_url_source == "uploads_legacy":
            print(f"使用旧路径网址文件：{website_url_path}")
        else:
            print(f"使用网址文件：{website_url_path}")

        website_data = load_website_json(
            project_name
        )

        website_inventory, page_map = (
            build_page_inventory(
                website_data=website_data,
                max_pages=max_page_candidates,
                excerpt_chars=excerpt_chars,
            )
        )

        print(
            f"可提供给AI的网页页面数：{len(page_map)}"
        )

        if len(page_map) < 8:
            print(
                "提示：成功读取的页面少于8个，"
                "部分Sitelink可能需要复用页面。"
            )

        prompt = build_prompt(
            project_brief=project_brief,
            website_inventory=website_inventory,
        )
        sitelinks_status = "generated"
    else:
        print(
            "客户资料未提供网站，本次跳过Sitelink，仅生成Callout。"
        )
        page_map = {}
        prompt = build_callout_only_prompt(
            project_brief=project_brief,
        )
        sitelinks_status = "skipped_no_website"

    result = call_llm(
        prompt=prompt,
        provider=provider,
        model=model,
        temperature=temperature,
        max_tokens=max_tokens,
        use_premium=use_premium,
    )

    content = result["content"]

    raw_path = (
        output_dir / "sitelink_callouts_raw.md"
    )

    raw_path.write_text(
        content,
        encoding="utf-8",
    )

    payload = extract_json_from_response(
        content
    )

    if not has_website:
        payload.setdefault("sitelinks", [])

    validation_errors = validate_payload(
        payload=payload,
        page_map=page_map,
        sitelinks_required=has_website,
    )

    if validation_errors:
        error_text = "\n".join(
            f"- {error}"
            for error in validation_errors
        )

        raise ValueError(
            "AI生成结果未通过校验：\n"
            f"{error_text}\n\n"
            f"AI原始输出已保存：{raw_path}"
        )

    enriched_data = enrich_payload(
        payload=payload,
        page_map=page_map,
        result_meta=result,
        sitelinks_status=sitelinks_status,
    )

    json_path = (
        output_dir / "sitelink_callouts.json"
    )
    xlsx_path = (
        output_dir / "sitelink_callouts.xlsx"
    )

    write_json(json_path, enriched_data)
    write_xlsx(xlsx_path, enriched_data)

    print("\nSitelink和Callout生成完成")
    print(f"Excel结果：{xlsx_path}")
    print(f"结构化JSON：{json_path}")
    print(f"AI原始输出：{raw_path}")

    return enriched_data


def main():
    parser = argparse.ArgumentParser(
        description=(
            "根据项目资料和已抓取网页生成"
            "Yandex Sitelink与Callout。"
        )
    )

    parser.add_argument(
        "--project",
        help="项目名称",
    )
    parser.add_argument(
        "--model",
        help="指定模型名称",
    )
    parser.add_argument(
        "--temperature",
        type=float,
        default=0.5,
    )
    parser.add_argument(
        "--max-tokens",
        type=int,
        default=5000,
    )
    parser.add_argument(
        "--use-premium",
        action="store_true",
        default=None,
    )
    parser.add_argument(
        "--max-page-candidates",
        type=int,
        default=25,
        help="最多发送给AI多少个网页候选，默认25",
    )
    parser.add_argument(
        "--excerpt-chars",
        type=int,
        default=1200,
        help="每个页面发送给AI的正文摘要字符数，默认1200",
    )

    args = parser.parse_args()

    project_name = (
        args.project
        or input("请输入项目名称：").strip()
    )

    if not project_name:
        raise ValueError("项目名称不能为空。")

    generate_sitelink_callouts(
        project_name=project_name,
        model=args.model,
        temperature=args.temperature,
        max_tokens=args.max_tokens,
        use_premium=args.use_premium,
        max_page_candidates=args.max_page_candidates,
        excerpt_chars=args.excerpt_chars,
    )


if __name__ == "__main__":
    main()
